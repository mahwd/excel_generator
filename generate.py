from datetime import datetime
import tkinter as tk
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.worksheet.table import Table, TableStyleInfo



window = tk.Tk()

window.geometry("500x500")

window.title("Məlumatları daxil edin") 

name=None

form_frame = tk.Frame(window)
form_frame.columnconfigure(0, weight=10)
form_frame.columnconfigure(1, weight=10)

muqavile_no_textbox = tk.Entry(form_frame, font="18")
muqavile_no_label = tk.Label(form_frame, font='18', text="Müqavilə №")
muqavile_no_label.grid(row=0, column=0, sticky='e')
muqavile_no_textbox.grid(row=0, column=1, sticky='w', padx=20)


buyer_fullname_textbox = tk.Entry(form_frame, font="18")
buyer_fullname_label = tk.Label(form_frame, font='18', text="Alıcının tam adı")
buyer_fullname_label.grid(row=1, column=0, sticky='e')
buyer_fullname_textbox.grid(row=1, column=1, sticky='w', padx=20)

ilkin_odenis_textbox = tk.Entry(form_frame, font="18")
ilkin_odenis_label = tk.Label(form_frame, font='18', text="İlkin ödəniş")
ilkin_odenis_label.grid(row=2, column=0, sticky='e')
ilkin_odenis_textbox.grid(row=2, column=1,sticky='w',padx=20)

mebleg_textbox = tk.Entry(form_frame, font="18")
mebleg_label = tk.Label(form_frame, font='18', text="Məbləğ")
mebleg_label.grid(row=3, column=0, sticky='e')
mebleg_textbox.grid(row=3, column=1,sticky='w',padx=20)

def export_excel():
    buyer_fullname = buyer_fullname_textbox.get()
    ilkin_odenis = ilkin_odenis_textbox.get()
    mebleg = mebleg_textbox.get()
    muqavile_no = muqavile_no_textbox.get()
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    ws['B2'] = '{}-cü il tarixli {} saylı Müddətli (nisyə) alqı-satqı müqaviləsinə Əlavə No1'.format(datetime.now().strftime('%d %B,%Y'), muqavile_no)
    ws['B2'].font = Font(name='Times New Roman', size=9, bold=True)

    ws['F4'] = '{} / Bakı şəhəri'.format(datetime.now().strftime('%d.%m.%Y'))

    ws['B5'] = '1. Mal haqqında məlumat:'


    # Save the file
    wb.save("new_sample.xlsx")


    window.quit()

generate_btn = tk.Button(form_frame, text='EXPORT EXCEL', command=export_excel)
generate_btn.grid(row=4, columnspan=2, sticky='NS')
form_frame.pack(fill='x')
window.mainloop()




