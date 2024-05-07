from datetime import datetime
import tkinter as tk
from openpyxl import Workbook, load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.worksheet.table import Table, TableStyleInfo
from dateutil import relativedelta


window = tk.Tk()

window.geometry("600x600")

window.title("Məlumatları daxil edin") 

name=None

form_frame = tk.Frame(window, pady=15)
form_frame.columnconfigure(0, weight=1)
form_frame.columnconfigure(1, weight=1)

muqavile_no_textbox = tk.Entry(form_frame, font="18")
muqavile_no_label = tk.Label(form_frame, font='18', text="Müqavilə №")
muqavile_no_label.grid(row=0, column=0, sticky='e')
muqavile_no_textbox.grid(row=0, column=1, sticky='w', padx=20)


alici_textbox = tk.Entry(form_frame, font="18")
alici_label = tk.Label(form_frame, font='18', text="Alıcının tam adı:")
alici_label.grid(row=1, column=0, sticky='e')
alici_textbox.grid(row=1, column=1, sticky='w', padx=20)

id_number_textbox = tk.Entry(form_frame, font="18")
id_number_label = tk.Label(form_frame, font='18', text="ŞV-nin Seriya Nömrəsi:")
id_number_label.grid(row=2, column=0, sticky='e')
id_number_textbox.grid(row=2, column=1, sticky='w', padx=20)

pincode_textbox = tk.Entry(form_frame, font="18")
pincode_label = tk.Label(form_frame, font='18', text="ŞV-nin Fin Kodu:")
pincode_label.grid(row=3, column=0, sticky='e')
pincode_textbox.grid(row=3, column=1, sticky='w', padx=20)

place_birth_textbox = tk.Entry(form_frame, font="18")
place_birth_label = tk.Label(form_frame, font='18', text="Doğulduğu yer və tarix:")
place_birth_label.grid(row=4, column=0, sticky='e')
place_birth_textbox.grid(row=4, column=1, sticky='w', padx=20)

expiration_date_textbox = tk.Entry(form_frame, font="18")
expiration_date_label = tk.Label(form_frame, font='18', text="ŞV-nin Bitmə Tarixi:")
expiration_date_label.grid(row=5, column=0, sticky='e')
expiration_date_textbox.grid(row=5, column=1, sticky='w', padx=20)

work_textbox = tk.Entry(form_frame, font="18")
work_label = tk.Label(form_frame, font='18', text="İş Yerləri:")
work_label.grid(row=6, column=0, sticky='e')
work_textbox.grid(row=6, column=1, sticky='w', padx=20)

registration_address_textbox = tk.Entry(form_frame, font="18")
registration_address_label = tk.Label(form_frame, font='18', text="Qeydiyyat Ünvanı:")
registration_address_label.grid(row=7, column=0, sticky='e')
registration_address_textbox.grid(row=7, column=1, sticky='w', padx=20)

resident_address_textbox = tk.Entry(form_frame, font="18")
resident_address_label = tk.Label(form_frame, font='18', text="Yaşayış Ünvanı:")
resident_address_label.grid(row=8, column=0, sticky='e')
resident_address_textbox.grid(row=8, column=1, sticky='w', padx=20)

ilkin_odenis_textbox = tk.Entry(form_frame, font="18")
ilkin_odenis_label = tk.Label(form_frame, font='18', text="İlkin ödəniş:")
ilkin_odenis_label.grid(row=9, column=0, sticky='e')
ilkin_odenis_textbox.grid(row=9, column=1,sticky='w',padx=20)

miqdar_textbox = tk.Entry(form_frame, font="18")
miqdar_label = tk.Label(form_frame, font='18', text="Miqdar:")
miqdar_label.grid(row=10, column=0, sticky='e')
miqdar_textbox.grid(row=10, column=1, sticky='w', padx=20)

ay_textbox = tk.Entry(form_frame, font="18")
ay_label = tk.Label(form_frame, font='18', text="Ay:")
ay_label.grid(row=11, column=0, sticky='e')
ay_textbox.grid(row=11, column=1, sticky='w', padx=20)

imei_textbox = tk.Entry(form_frame, font="18")
imei_label = tk.Label(form_frame, font='18', text="IMEI:")
imei_label.grid(row=12, column=0, sticky='e')
imei_textbox.grid(row=12, column=1,sticky='w',padx=20)

mebleg_textbox = tk.Entry(form_frame, font="18")
mebleg_label = tk.Label(form_frame, font='18', text="Məbləğ:")
mebleg_label.grid(row=13, column=0, sticky='e')
mebleg_textbox.grid(row=13, column=1,sticky='w',padx=20)


nisye_mebleg_textbox = tk.Entry(form_frame, font="18")
nisye_mebleg_label = tk.Label(form_frame, font='18', text="Nisyə məbləğ:")
nisye_mebleg_label.grid(row=14, column=0, sticky='e')
nisye_mebleg_textbox.grid(row=14, column=1,sticky='w',padx=20)

mehsul_textbox = tk.Text(form_frame, font="18", height=3, width=21)
mehsul_label = tk.Label(form_frame, font='18', text="Məhsul adı:")
mehsul_label.grid(row=15, column=0, sticky='e')
mehsul_textbox.grid(row=15, column=1,sticky='w',padx=20)

contacts_textbox = tk.Text(form_frame, font="18", height=4, width=21)
contacts_label = tk.Label(form_frame, font='18', text="Əlaqələr:")
contacts_label.grid(row=16, column=0, sticky='e')
contacts_textbox.grid(row=16, column=1,sticky='w',padx=20)



# insert temp data
muqavile_no_textbox.insert(0, "015")
alici_textbox.insert(0, "Tagiyev Hesen Ilham oglu")
id_number_textbox.insert(0, "AA2282475")
pincode_textbox.insert(0, "5p1155r")
place_birth_textbox.insert(0, "Baki/09.10.93")
expiration_date_textbox.insert(0, "02.02.2029")
work_textbox.insert(0, "MDM")
registration_address_textbox.insert(0, "Baki")
resident_address_textbox.insert(0, "Sumqayit")
contacts_textbox.insert(1.0, "Rasim,qohum,0704454884\nRevan,dost,0505643876")
ilkin_odenis_textbox.insert(0, "200")
miqdar_textbox.insert(0, "2000")
ay_textbox.insert(0, "12")
imei_textbox.insert(0, "123455678")
mebleg_textbox.insert(0, "2200")
nisye_mebleg_textbox.insert(0, "2200")
mehsul_textbox.insert(1.0, "Xiaomi 14 Ultra 16 GB/ 512 GB Black")

def export_excel():
    today = datetime.now()

    muqavile_no = muqavile_no_textbox.get()
    alici = alici_textbox.get()
    id_number = id_number_textbox.get()
    pincode = pincode_textbox.get()
    place_birth = place_birth_textbox.get()
    expiration_date = expiration_date_textbox.get()
    work = work_textbox.get()
    registration_address = registration_address_textbox.get()
    resident_address = resident_address_textbox.get()
    contacts = contacts_textbox.get("1.0", "end-1c")
    ilkin_odenis = ilkin_odenis_textbox.get()
    miqdar = miqdar_textbox.get()
    ay = ay_textbox.get()
    imei = imei_textbox.get()
    mebleg = mebleg_textbox.get()
    nisye_mebleg = nisye_mebleg_textbox.get()
    mehsul = mehsul_textbox.get("1.0", "end-1c")


    wb = load_workbook('./samples/yeni_muqavile_18ay.xlsx') if ay == 18 else load_workbook('./samples/yeni_muqavile_12ay.xlsx')

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    ws['A2'] = '{}-cü il tarixli {} saylı Müddətli (nisyə) alqı-satqı müqaviləsinə Əlavə No1'.format(today.strftime('%d %B,%Y'), muqavile_no)

    ws['E4'] = '{} / Bakı şəhəri'.format(today.strftime('%d.%m.%Y'))

    # table 1
    ws['A8'] = 1
    ws['B8'] = mehsul
    ws['C8'] = mebleg
    ws['D8'] = miqdar
    ws['E8'] = nisye_mebleg
    ws['F8'] = imei
    
    # 2. Nisyə alqı satqı müqaviləsinin ödəniş cədvəli: table 1
    ws['C17'] = alici
    ws['C18'] = ilkin_odenis
    ws['C19'] = nisye_mebleg
    ws['C21'] = ay

    # table 2
    nextmonth = today
    for i in range(int(ay)):
        date_index = "B" + str(26+i)
        index = "C"+str(26+i)
        ws[date_index] = nextmonth.strftime('%d.%m.%Y')
        ws[index] = i+1
        nextmonth = nextmonth + relativedelta.relativedelta(months=1)

    # Alıcı məlumatları table 1
    ws['C71'] = alici
    ws['C72'] = id_number
    ws['C73'] = place_birth
    ws['C74'] = pincode
    ws['C75'] = expiration_date
    ws['C76'] = work
    ws['C77'] = registration_address
    ws['C78'] = resident_address

    try:
        for x, line in enumerate(contacts_textbox.get('1.0', 'end-1c').splitlines()):
            row = str(82 + x)
            # Iterate lines
            if line:
                [contact_name, contact_desc, contact_phone] = line.split(',')
                ws['B'+row] = contact_name
                ws['C'+row] = contact_desc
                ws['D'+row] = contact_phone

    except Exception as e:
        print(e)


    ws["A89"] = '{} \
                     \n                                                      (imza)________'.format(alici)
    ws['E100'] = '{} / Bakı şəhəri'.format(today.strftime('%d.%m.%Y'))
    ws["A102"] = 'Bir tərəfdən VÖEN:1503125972 əsasında fəaliyyət göstərən “MOBİOS ELECTRONİCS”,  mağazasının müdiri T.Lütvəlizadə şəxsində bundan sonra “Satıcı” adlanacaq və digər tərəfdən {}, bundan sonra “Alıcı” adlanacaq, aşağıdakı şərtlər əsasında həmin müqaviləni (bundan sonra “Müqavilə”) bağladılar.'.format(alici)
    
    # Save the file
    wb.save("new_sample.xlsx")


    window.quit()

generate_btn = tk.Button(form_frame, text='EXPORT EXCEL', command=export_excel)
generate_btn.grid(row=17, columnspan=2, sticky='NS')
form_frame.pack(fill='x')
window.mainloop()




